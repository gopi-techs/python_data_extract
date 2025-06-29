import pandas as pd
import os
from datetime import datetime
import re
from openpyxl import load_workbook
import warnings

warnings.filterwarnings('ignore')


def extract_invoice_data(file_path):
    """
    Extract key invoice data from Excel file based on the observed structure.
    """
    try:
        # Load workbook to get sheet names
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        # Try to find the main data sheet (skip 'book' sheet, prefer date-named sheets)
        target_sheet = None
        for sheet in sheet_names:
            if sheet != 'book' and len(sheet) > 3:  # Skip 'book' and very short names
                target_sheet = sheet
                break

        if not target_sheet:
            target_sheet = sheet_names[0]  # Fallback to first sheet

        # Read the Excel file
        df = pd.read_excel(file_path, sheet_name=target_sheet, header=None)

        # Initialize result dictionary
        result = {
            'file_name': os.path.basename(file_path),
            'sheet_name': target_sheet,
            'invoice_number': '',
            'invoice_date': '',
            'company_name': '',
            'vendor_code': '',
            'po_number': '',
            'po_date': '',
            'gstin': '',
            'total_value': '',
            'gross_total_after_tax': '',
            'extraction_status': 'Success'
        }

        # Convert all cells to string for easier searching
        df = df.astype(str).fillna('')

        # Search for key data points in the structured format
        for index, row in df.iterrows():
            row_values = row.tolist()

            # Look for invoice number (SSH pattern)
            if len(row_values) > 5:
                cell_value = str(row_values[5]).strip()
                if 'SSH' in cell_value and '/' in cell_value:
                    result['invoice_number'] = cell_value

                # Look for dates in column 5
                if re.match(r'\d{4}-\d{2}-\d{2}', cell_value):
                    if 'invoice_date' not in result or not result['invoice_date']:
                        result['invoice_date'] = cell_value.split(' ')[0]  # Remove time part

            # Search through all columns for financial totals
            for col_idx, cell_value in enumerate(row_values):
                cell_str = str(cell_value).strip()

                # Look for GROSS TOTAL AFTER TAX
                if 'GROSS TOTAL AFTER' in cell_str.upper() and 'TAX' in cell_str.upper():
                    # Look for the value in adjacent columns
                    for next_col in range(col_idx + 1, min(len(row_values), col_idx + 4)):
                        next_value = str(row_values[next_col]).strip()
                        if re.match(r'[\d,]+\.?\d*', next_value) and len(next_value) > 2:
                            try:
                                # Validate it's a number
                                float(next_value.replace(',', ''))
                                result['gross_total_after_tax'] = next_value
                                break
                            except (ValueError, TypeError):
                                continue

                # Look for various TOTAL patterns (more comprehensive)
                total_keywords = ['TOTAL VALUE', 'TOTAL AMOUNT', 'TOTAL', 'AMOUNT', 'VALUE']
                for keyword in total_keywords:
                    if keyword in cell_str.upper() and 'GROSS' not in cell_str.upper():
                        # Look for the value in adjacent columns
                        for next_col in range(col_idx + 1, min(len(row_values), col_idx + 4)):
                            next_value = str(row_values[next_col]).strip()
                            if re.match(r'[\d,]+\.?\d*', next_value) and len(next_value) > 2:
                                try:
                                    # Validate it's a number
                                    num_val = float(next_value.replace(',', ''))
                                    if num_val > 100 and not result['total_value']:  # Take the first significant one found
                                        result['total_value'] = next_value
                                        break
                                except (ValueError, TypeError):
                                    continue
                        if result['total_value']:  # Break outer loop if found
                            break

                # Alternative pattern: Look for numerical values in rightmost columns
                if col_idx >= len(row_values) - 3:  # Last three columns (expanded search)
                    # Improved number matching with better error handling
                    if re.match(r'[\d,]+\.?\d*', cell_str) and len(cell_str) > 2:
                        try:
                            # Try to convert to float to validate it's a number
                            num_value = float(cell_str.replace(',', ''))
                            if num_value > 100:
                                # Check if the row contains relevant keywords
                                row_text = ' '.join(row_values).upper()
                                if 'GROSS' in row_text and 'TOTAL' in row_text and not result['gross_total_after_tax']:
                                    result['gross_total_after_tax'] = cell_str
                                elif any(keyword in row_text for keyword in ['TOTAL', 'AMOUNT', 'VALUE']) and not result['total_value']:
                                    # More flexible matching for total value
                                    result['total_value'] = cell_str
                        except (ValueError, TypeError):
                            # Skip if conversion fails
                            continue

        # Post-processing: If we have gross_total_after_tax but no total_value,
        # try to extract base amount from gross total
        if result['gross_total_after_tax'] and not result['total_value']:
            try:
                gross_val = float(result['gross_total_after_tax'].replace(',', ''))
                # Estimate base value (assuming ~18% tax rate)
                estimated_base = gross_val / 1.18
                # Don't set this as total_value as it's estimated, but we could add a note
                pass
            except:
                pass

            # Look for company name (following "Name :" pattern)
            if len(row_values) > 0:
                cell_value = str(row_values[0]).strip()
                if cell_value.startswith('Name  :'):
                    company_name = cell_value.replace('Name  :', '').strip()
                    if company_name:
                        result['company_name'] = company_name

                # Look for GSTIN
                if cell_value.startswith('GSTIN :'):
                    gstin = cell_value.replace('GSTIN :', '').strip()
                    if gstin and len(gstin) > 10:  # Valid GSTIN should be longer
                        result['gstin'] = gstin

            # Look for vendor code and PO details in column 5
            if len(row_values) > 5:
                if len(row_values) > 3:
                    label = str(row_values[3]).strip()
                    value = str(row_values[5]).strip()

                    if 'Vendor Code' in label and value:
                        result['vendor_code'] = value
                    elif 'PO Date' in label and value:
                        result['po_number'] = value  # This seems to be PO number based on debug
                    elif 'Purchase Order No' in label and value:  # Note the typo in original
                        if re.match(r'\d{4}-\d{2}-\d{2}', value):
                            result['po_date'] = value.split(' ')[0]

        return result

    except Exception as e:
        return {
            'file_name': os.path.basename(file_path),
            'sheet_name': 'Error',
            'invoice_number': '',
            'invoice_date': '',
            'company_name': '',
            'vendor_code': '',
            'po_number': '',
            'po_date': '',
            'gstin': '',
            'total_value': '',
            'gross_total_after_tax': '',
            'extraction_status': f'Error: {str(e)}'
        }


def process_all_invoices(directory_path):
    """
    Process all Excel files in the directory and extract invoice data.
    """
    # Find all Excel files
    excel_files = []
    for root, dirs, files in os.walk(directory_path):
        for file in files:
            if file.endswith(('.xlsx', '.xls')):
                excel_files.append(os.path.join(root, file))

    print(f"🔍 Found {len(excel_files)} Excel files")

    # Process each file
    all_results = []
    successful_extractions = 0

    for i, file_path in enumerate(excel_files, 1):
        print(f"📄 Processing {i}/{len(excel_files)}: {os.path.basename(file_path)}")

        result = extract_invoice_data(file_path)
        all_results.append(result)

        if result['extraction_status'] == 'Success':
            successful_extractions += 1
            # Print key extracted data for verification
            if result['invoice_number']:
                print(f"   ✅ Invoice: {result['invoice_number']}")
            if result['company_name']:
                print(f"   ✅ Company: {result['company_name']}")
            if result['invoice_date']:
                print(f"   ✅ Date: {result['invoice_date']}")
        else:
            print(f"   ❌ {result['extraction_status']}")

    # Create DataFrame and save to Excel
    df_results = pd.DataFrame(all_results)

    # Generate output filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"invoice_data_extracted_{timestamp}.xlsx"

    # Save results
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_results.to_excel(writer, sheet_name='Extracted_Data', index=False)

        # Create a summary sheet
        summary_data = {
            'Metric': [
                'Total Files Processed',
                'Successful Extractions',
                'Failed Extractions',
                'Success Rate (%)',
                'Files with Invoice Numbers',
                'Files with Company Names',
                'Files with Dates',
                'Files with Total Values',
                'Files with Gross Total After Tax'
            ],
            'Count': [
                len(excel_files),
                successful_extractions,
                len(excel_files) - successful_extractions,
                round((successful_extractions / len(excel_files)) * 100, 2) if excel_files else 0,
                len(df_results[df_results['invoice_number'] != '']),
                len(df_results[df_results['company_name'] != '']),
                len(df_results[df_results['invoice_date'] != '']),
                len(df_results[df_results['total_value'] != '']),
                len(df_results[df_results['gross_total_after_tax'] != ''])
            ]
        }

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

    print(f"\n📊 EXTRACTION COMPLETE!")
    print(f"📁 Results saved to: {output_file}")
    print(f"✅ Successfully processed: {successful_extractions}/{len(excel_files)} files")
    print(f"📈 Success rate: {(successful_extractions / len(excel_files) * 100):.1f}%")

    return df_results


def main():
    # Set the directory path containing Excel files
    directory_path = r"F:\Office\oldinvoicesbkp\SRI SAI HEATERS APRIL 2024 - MAR 2025"  # Update this path as needed

    print("🚀 Starting Invoice Data Extraction...")
    print(f"📂 Directory: {directory_path}")

    if not os.path.exists(directory_path):
        print(f"❌ Directory not found: {directory_path}")
        return

    # Process all invoices
    results_df = process_all_invoices(directory_path)

    # Display sample results
    print("\n📋 SAMPLE EXTRACTED DATA:")
    print("=" * 80)
    sample_data = results_df[results_df['extraction_status'] == 'Success'].head(5)

    for _, row in sample_data.iterrows():
        print(f"File: {row['file_name']}")
        print(f"  Invoice: {row['invoice_number']}")
        print(f"  Company: {row['company_name']}")
        print(f"  Date: {row['invoice_date']}")
        print(f"  Vendor Code: {row['vendor_code']}")
        print(f"  Total Value: {row['total_value']}")
        print(f"  Gross Total After Tax: {row['gross_total_after_tax']}")
        print("-" * 40)


if __name__ == "__main__":
    main()